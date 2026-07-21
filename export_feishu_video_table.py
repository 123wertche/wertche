"""Export every field and record from the configured Feishu video table to JSON."""

import argparse
import json
from datetime import datetime
from pathlib import Path

import download_bili_following_latest as bili


ROOT = Path(__file__).resolve().parent
TARGET_VIDEO_TABLE_ID = "tblakZnkghpokyGT"


def ordered_record(record, field_names):
    return {"record_id": record.get("_record_id"), **{name: record.get(name) for name in field_names}}


def _excel_value(value):
    if value is None or isinstance(value, (int, float, bool)):
        return value
    if isinstance(value, str):
        return "'" + value if value.startswith("=") else value
    return json.dumps(value, ensure_ascii=False, separators=(",", ":"))


def write_xlsx(path, payload):
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.worksheet.table import Table, TableStyleInfo

    path = Path(path)
    path.parent.mkdir(parents=True, exist_ok=True)
    book = Workbook()
    sheet = book.active
    sheet.title = "视频表"
    fields = [item["name"] for item in payload.get("fields", [])]
    headers = ["record_id", *fields]
    sheet.append(headers)
    for record in payload.get("records", []):
        sheet.append([_excel_value(record.get(header)) for header in headers])
    # Preserve complete cell values while preventing long transcripts from
    # expanding a rendered row into a multi-page-height canvas.
    for row_index in range(2, sheet.max_row + 1):
        sheet.row_dimensions[row_index].height = 45
    header_fill = PatternFill("solid", fgColor="0F2747")
    for cell in sheet[1]:
        cell.fill = header_fill
        cell.font = Font(color="FFFFFF", bold=True)
        cell.alignment = Alignment(vertical="center", wrap_text=True)
    sheet.freeze_panes = "C2"
    sheet.auto_filter.ref = sheet.dimensions
    if sheet.max_row > 1:
        table = Table(displayName="FeishuVideoTable", ref=sheet.dimensions)
        table.tableStyleInfo = TableStyleInfo(name="TableStyleMedium2", showRowStripes=True, showColumnStripes=False)
        sheet.add_table(table)
    for index, header in enumerate(headers, 1):
        descriptive = any(token in header for token in ("标题", "摘要", "要点", "说明", "路径", "链接", "评论", "文案"))
        sheet.column_dimensions[sheet.cell(1, index).column_letter].width = 34 if descriptive else min(24, max(13, len(header) * 2 + 3))
        for cell in sheet.iter_cols(min_col=index, max_col=index, min_row=2, max_row=sheet.max_row):
            for item in cell:
                item.alignment = Alignment(vertical="top", wrap_text=descriptive)
                if any(token in header for token in ("播放量", "点赞数", "评论数", "转发数", "收藏数", "时长秒", "粉丝数")):
                    item.number_format = "#,##0"
                elif any(token in header for token in ("完播率", "跳出率")):
                    item.number_format = "0.00%"

    dictionary = book.create_sheet("字段说明")
    dictionary.append(["字段名称", "字段 ID", "字段类型"])
    for item in payload.get("fields", []):
        dictionary.append([item.get("name"), item.get("field_id"), _excel_value(item.get("type"))])
    for cell in dictionary[1]:
        cell.fill = header_fill
        cell.font = Font(color="FFFFFF", bold=True)
    dictionary.freeze_panes = "A2"
    dictionary.column_dimensions["A"].width = 28
    dictionary.column_dimensions["B"].width = 24
    dictionary.column_dimensions["C"].width = 18
    if dictionary.max_row > 1:
        table = Table(displayName="FeishuFieldDictionary", ref=dictionary.dimensions)
        table.tableStyleInfo = TableStyleInfo(name="TableStyleMedium2", showRowStripes=True, showColumnStripes=False)
        dictionary.add_table(table)

    summary = book.create_sheet("导出说明", 0)
    summary.append(["飞书视频表完整导出", ""])
    summary.merge_cells("A1:B1")
    summary["A1"].fill = header_fill
    summary["A1"].font = Font(color="FFFFFF", bold=True, size=16)
    for row in (
        ("导出时间", payload.get("exported_at")),
        ("飞书多维表", payload.get("base_name")),
        ("目标表 ID", (payload.get("table") or {}).get("table_id")),
        ("字段数量", len(fields)),
        ("记录数量", len(payload.get("records", []))),
    ):
        summary.append(row)
    for cell in summary["A"][1:]:
        cell.fill = PatternFill("solid", fgColor="E8F0FA")
        cell.font = Font(color="0F2747", bold=True)
    summary.column_dimensions["A"].width = 18
    summary.column_dimensions["B"].width = 60
    book.save(path)


def main():
    parser = argparse.ArgumentParser(allow_abbrev=False)
    parser.add_argument("--output", required=True)
    parser.add_argument("--xlsx-output")
    args = parser.parse_args()
    output = Path(args.output)
    if not output.is_absolute():
        output = ROOT / output
    if output.exists():
        raise FileExistsError(f"refusing to overwrite existing export: {output}")

    config = bili.load_config()
    table = config["tables"]["videos"]
    if table["table_id"] != TARGET_VIDEO_TABLE_ID:
        raise RuntimeError("configured video table does not match the required target")
    fields_by_name = bili.field_names(config, table["table_id"])
    field_names = list(fields_by_name)
    rows = bili.list_records(config, table["table_id"], field_names)
    payload = {
        "exported_at": datetime.now().astimezone().isoformat(timespec="seconds"),
        "base_name": config.get("base_name"),
        "base_url": config.get("base_url"),
        "table": table,
        "field_count": len(field_names),
        "record_count": len(rows),
        "fields": [
            {
                "name": name,
                "field_id": spec.get("field_id") or spec.get("id"),
                "type": spec.get("type"),
            }
            for name, spec in fields_by_name.items()
        ],
        "records": [ordered_record(row, field_names) for row in rows],
    }
    output.parent.mkdir(parents=True, exist_ok=True)
    output.write_text(json.dumps(payload, ensure_ascii=False, indent=2) + "\n", encoding="utf-8")
    xlsx_output = None
    if args.xlsx_output:
        xlsx_output = Path(args.xlsx_output)
        if not xlsx_output.is_absolute():
            xlsx_output = ROOT / xlsx_output
        if xlsx_output.exists():
            raise FileExistsError(f"refusing to overwrite existing export: {xlsx_output}")
        write_xlsx(xlsx_output, payload)
    print(
        json.dumps(
            {
                "output": str(output),
                "table_id": table["table_id"],
                "field_count": len(field_names),
                "record_count": len(rows),
                "xlsx_output": str(xlsx_output) if xlsx_output else None,
            },
            ensure_ascii=False,
        )
    )


if __name__ == "__main__":
    main()
