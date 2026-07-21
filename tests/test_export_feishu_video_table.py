import tempfile
import unittest
from pathlib import Path

from openpyxl import load_workbook

from export_feishu_video_table import ordered_record, write_xlsx


class ExportRecordTests(unittest.TestCase):
    def test_xlsx_export_reopens_with_headers_and_complex_cells(self):
        payload = {
            "exported_at": "2026-07-20T12:00:00+08:00",
            "base_name": "Base",
            "table": {"table_id": "tbl"},
            "fields": [
                {"name": "播放量", "field_id": "fld1", "type": "number"},
                {"name": "视频封面", "field_id": "fld2", "type": "attachment"},
            ],
            "records": [{"record_id": "rec1", "播放量": 0, "视频封面": [{"file_token": "token"}]}],
        }
        with tempfile.TemporaryDirectory() as directory:
            target = Path(directory) / "export.xlsx"

            write_xlsx(target, payload)

            book = load_workbook(target, read_only=True, data_only=False)
            self.assertEqual(book["视频表"]["A1"].value, "record_id")
            self.assertEqual(book["视频表"]["B2"].value, 0)
            self.assertIn("file_token", book["视频表"]["C2"].value)
            book.close()

    def test_keeps_all_fields_in_table_order_and_record_id(self):
        record = {"字段B": 2, "字段A": "a", "_record_id": "rec1"}

        exported = ordered_record(record, ["字段A", "字段B", "字段C"])

        self.assertEqual(
            exported,
            {"record_id": "rec1", "字段A": "a", "字段B": 2, "字段C": None},
        )


    def test_caps_data_row_height_for_long_text_exports(self):
        payload = {
            "exported_at": "2026-07-20T12:00:00+08:00",
            "base_name": "Base",
            "table": {"table_id": "tbl"},
            "fields": [{"name": "Long text", "field_id": "fld1", "type": "text"}],
            "records": [{"record_id": "rec1", "Long text": "content " * 2000}],
        }
        with tempfile.TemporaryDirectory() as directory:
            target = Path(directory) / "export.xlsx"

            write_xlsx(target, payload)

            book = load_workbook(target, read_only=False, data_only=False)
            self.assertLessEqual(book.worksheets[1].row_dimensions[2].height, 60)
            book.close()


if __name__ == "__main__":
    unittest.main()
